import streamlit as st
import pdfplumber
import pandas as pd
import re
from io import BytesIO
from openpyxl.styles import Font
from openpyxl import load_workbook

st.title("Conversor LRCO: PDF ➡️ Excel com Disciplinas Validadas 📄➡️📊")

uploaded_files = st.file_uploader("📥 Selecione os arquivos PDF do relatório LRCO", type="pdf", accept_multiple_files=True)
disciplinas_file = st.file_uploader("📚 Envie a planilha com a lista oficial de disciplinas", type=["xlsx"])

if uploaded_files and disciplinas_file:
    # Carrega lista de disciplinas válidas
    disciplinas_validas_df = pd.read_excel(disciplinas_file)
    disciplinas_validas = [d.strip().upper() for d in disciplinas_validas_df.iloc[:, 0].dropna().unique()]

    dados = []
    horario_re = r"\d{2}:\d{2}:\d{2}"
    registro_re = r"\d{2}/\d{2}/\d{4} \d{2}:\d{2}:\d{2}"
    data_relatorio_re = r"\b\d{2}/\d{2}/\d{4}\b"

    for uploaded_file in uploaded_files:
        turma_atual = None
        nome_escola = "ESCOLA NÃO IDENTIFICADA"
        municipio = "MUNICÍPIO NÃO IDENTIFICADO"
        data_relatorio = "DATA NÃO IDENTIFICADA"

        with pdfplumber.open(uploaded_file) as pdf:
            for page_num, page in enumerate(pdf.pages):
                texto = page.extract_text()
                if not texto:
                    continue
                linhas = texto.split("\n")

                if page_num == 0:
                    for i, linha in enumerate(linhas):
                        if "ESTADO DO PARANÁ" in linha:
                            match_data = re.search(data_relatorio_re, linha)
                            if match_data:
                                data_relatorio = match_data.group()
                        if "SECRETARIA DE ESTADO DA EDUCAÇÃO" in linha:
                            municipio = linha.split("SECRETARIA")[0].strip()
                            if i + 1 < len(linhas):
                                nome_escola = linhas[i + 1].strip()

                for linha in linhas:
                    linha = linha.strip()
                    if " - " in linha and "TURMA" not in linha and "LANÇAMENTO" not in linha:
                        turma_atual = linha
                        continue
                    if not turma_atual:
                        continue

                    horarios = re.findall(horario_re, linha)
                    registros = re.findall(registro_re, linha)

                    if not horarios:
                        continue

                    horario = horarios[0]
                    pos_horario = linha.find(horario)
                    pos_fim_horario = pos_horario + len(horario)

                    registro_aula = registros[0] if len(registros) >= 1 else "Sem registro"
                    registro_conteudo = registros[1] if len(registros) >= 2 else "Sem registro"

                    pos_registro = linha.find(registros[0]) if registros else len(linha)
                    disciplina_raw = linha[pos_fim_horario:pos_registro].strip()

                    # Validação da disciplina
                    disciplina_encontrada = None
                    for nome_disciplina in disciplinas_validas:
                        if nome_disciplina in disciplina_raw.upper():
                            disciplina_encontrada = nome_disciplina
                            break

                    if not disciplina_encontrada:
                        continue  # pula linha se disciplina não reconhecida

                    dados.append([
                        data_relatorio,
                        municipio,
                        nome_escola,
                        turma_atual,
                        horario,
                        disciplina_encontrada,
                        registro_aula,
                        registro_conteudo
                    ])

    colunas = [
        "DATA DO RELATÓRIO", "MUNICÍPIO", "ESCOLA", "TURMA",
        "HORÁRIO", "DISCIPLINA", "REGISTRO DE AULA", "REGISTRO DE CONTEÚDO"
    ]
    df = pd.DataFrame(dados, columns=colunas)

    st.success("✅ Conversão concluída! Veja a prévia abaixo.")
    st.dataframe(df)

    # Gera Excel com destaque
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Relatório")
        ws = writer.sheets["Relatório"]

        red_font = Font(color="FF0000")

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=7, max_col=8):
            for cell in row:
                if cell.value == "Sem registro":
                    cell.font = red_font

    output.seek(0)

    st.download_button(
        "📥 Baixar Excel Final",
        data=output,
        file_name="relatorio_validado.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
